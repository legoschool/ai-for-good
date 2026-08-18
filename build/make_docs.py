# -*- coding: utf-8 -*-
"""학교자율시간 진도표(XLSX) · 카드 교구 3종 · 교사용 해설서를 만든다.

사용법 : py -3 build/make_docs.py <진도표|교구|해설서|all>
"""
import copy
import io
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import hwpx
import tasks as T
from make_webapp import INFO_CARDS, SCENE_CARDS, SITUATION_CARDS

T.setup_console()

THIN = Side(style="thin", color="D0D5DD")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD = PatternFill("solid", fgColor="F3F4F6")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_sheet(ws, widths, header_row=1):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=header_row):
        for c in row:
            c.border = BOX
            c.alignment = WRAP
    for c in ws[header_row]:
        c.font = Font(bold=True)
        c.fill = HEAD
        c.alignment = CENTER
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def make_progress(data):
    wb = Workbook()

    ws = wb.active
    ws.title = "진도표"
    ws.append(["차시", "모듈", "학습 주제", "학습 문제", "성취기준", "교과·시수",
               "중점 휴먼스킬", "학생 산출물", "평가 방법", "웹앱", "운영 예정일"])
    for l in data["lessons"]:
        m = data["modules"][l["module"] - 1]
        ws.append([
            l["no"], "모듈%d %s" % (m["no"], m["name"]), l["shortTitle"], l["problem"],
            " ".join(l["standards"]), l["subject"],
            ", ".join(f["name"] for f in l["humanSkills"]["focus"]),
            ", ".join(l["outputs"]), ", ".join(l["assessment"]),
            l["webapp"]["name"], "",
        ])
    style_sheet(ws, [6, 16, 26, 44, 20, 22, 22, 32, 20, 22, 14])

    ws2 = wb.create_sheet("시수 배분")
    ws2.append(["교과·영역", "차시", "시수", "근거"])
    rows = [
        ["실과 디지털 사회와 인공지능", "1·2·3·4·9·11차시", 6, "[6실05-03] [6실05-04] [6실05-05]"],
        ["도덕 관계", "5·6차시", 2, "[6도02-03]"],
        ["국어 매체", "2차시 통합", 1, "[6국06-02]"],
        ["학교자율시간 또는 창의적 체험활동", "7·8·10·12차시", 4, "자율·자치 활동"],
    ]
    for r in rows:
        ws2.append(r)
    ws2.append(["합계", "", sum(r[2] for r in rows), "별도 시수 증가 없음"])
    style_sheet(ws2, [34, 24, 8, 40])

    ws3 = wb.create_sheet("평가 계획")
    ws3.append(["역량", "관련 차시", "평가 목표", "상", "중", "하", "평가 방법"])
    for p in data["assessmentPlan"]:
        ws3.append([p["skill"], ", ".join("%d차시" % n for n in p["lessons"]),
                    p["goal"], p["high"], p["mid"], p["low"], ", ".join(p["method"])])
    style_sheet(ws3, [16, 18, 34, 40, 30, 30, 20])

    ws4 = wb.create_sheet("사전사후 설문")
    ws4.append(["번호", "문항", "연계 휴먼스킬", "채점", "사전 평균", "사후 평균", "변화량"])
    s = data["survey"]
    for i in s["items"]:
        ws4.append([i["no"], i["text"], i["skill"], i["scoring"], None, None, None])
    first, last = 2, 1 + len(s["items"])
    for r in range(first, last + 1):
        ws4.cell(row=r, column=7).value = "=IF(AND(E%d<>\"\",F%d<>\"\"),F%d-E%d,\"\")" % (r, r, r, r)
    ws4.append(["", "전체 평균", "", "",
                "=IF(COUNT(E%d:E%d)=0,\"\",AVERAGE(E%d:E%d))" % (first, last, first, last),
                "=IF(COUNT(F%d:F%d)=0,\"\",AVERAGE(F%d:F%d))" % (first, last, first, last),
                "=IF(COUNT(G%d:G%d)=0,\"\",AVERAGE(G%d:G%d))" % (first, last, first, last)])
    style_sheet(ws4, [6, 50, 18, 10, 12, 12, 10])
    fill = PatternFill("solid", fgColor="FEF9C3")
    for r in range(first, last + 1):
        ws4.cell(row=r, column=5).fill = fill
        ws4.cell(row=r, column=6).fill = fill
    ws4["B%d" % (last + 3)] = "노란 칸에 학급 평균을 넣으면 변화량이 자동으로 계산됩니다."
    ws4["B%d" % (last + 4)] = s["note"]

    out = os.path.join(T.ROOT, "out", "서류", "WISE_학교자율시간_진도표.xlsx")
    d = os.path.dirname(out)
    if not os.path.isdir(d):
        os.makedirs(d)
    wb.save(out)
    print("만들었다 : %s  (시트 %d개)" % (out, len(wb.sheetnames)))
    return out


# ---------- HWPX 문서 만들기 (문단 복제 방식) ----------

def build_hwpx(lines, out_path):
    root = hwpx.load_section()
    kids = list(root)
    tmpl = {"title": kids[0], "body": kids[2], "blank": kids[1]}

    built = []
    for kind, value in lines:
        par = copy.deepcopy(tmpl[kind])
        hwpx.strip_linesegarray(par)
        nodes = list(par.iter(hwpx.P + "t"))
        if nodes:
            nodes[0].text = value
            for n in nodes[1:]:
                n.text = ""
        built.append(par)

    for k in kids:
        root.remove(k)
    for par in built:
        root.append(par)
    hwpx.strip_linesegarray(root)

    body = "\n".join(hwpx.paragraph_text(p) for p in root)
    hwpx.write_hwpx(out_path, hwpx.serialize(root), prv_text=body)
    return len(built)


def make_cards(data):
    sets = [
        ("정보 분류 카드", 4, INFO_CARDS,
         "넣어도 돼 / 조건을 지키면 돼 / 넣으면 안 돼 세 칸으로 나눕니다."),
        ("AI 활용 장면 카드", 5, SCENE_CARDS,
         "보조 / 대행 두 칸으로 나눕니다."),
        ("상황 카드", 6, SITUATION_CARDS,
         "초록 / 노랑 / 주황 / 빨강 네 칸으로 나눕니다."),
    ]
    made = []
    for name, lesson_no, cards, howto in sets:
        L = [("title", "%s  (%d차시)" % (name, lesson_no)), ("blank", "")]
        L.append(("body", "카드 %d장 · 학급당 1세트 · 두꺼운 종이에 인쇄해 잘라 씁니다." % len(cards)))
        L.append(("body", howto))
        L.append(("blank", ""))
        L.append(("body", "----------------------------------------------------------------------"))
        for i, c in enumerate(cards, 1):
            L.append(("body", ""))
            L.append(("body", "%2d.  %s" % (i, c)))
            L.append(("body", ""))
            L.append(("body", "----------------------------------------------------------------------"))
        L.append(("blank", ""))
        L.append(("body", "잘라 쓰는 법 : 가로선을 따라 자르면 카드 한 장이 됩니다."))
        L.append(("body", "웹앱 없이도 이 카드만으로 같은 활동을 할 수 있습니다."))
        L.append(("blank", ""))
        L.append(("body", data["program"]["copyrightLine"]))

        out = os.path.join(T.ROOT, "out", "교구", "WISE_%s.hwpx" % name.replace(" ", "_"))
        n = build_hwpx(L, out)
        print("만들었다 : %s  (카드 %d장, 문단 %d개)" % (out, len(cards), n))
        made.append(out)
    return made


def make_guide(data):
    prog = data["program"]
    L = []
    a = L.append

    a(("title", "%s  교사용 해설서" % prog["name"]))
    a(("blank", ""))
    a(("body", prog["team"]))
    a(("body", " · ".join(prog["members"])))
    a(("blank", ""))

    a(("body", "1. 이 프로그램은 무엇인가"))
    a(("body", "핵심 질문 : %s" % prog["subtitle"]))
    a(("body", "대상 : %s · 총 %d차시 · 적용 %s" % (prog["target"], prog["totalLessons"], prog["period"])))
    a(("body", "학습 경로 : 발견(모듈1) → 판단(모듈2) → 실천(모듈3)"))
    a(("body", "모든 차시에 %s의 흐름을 넣었습니다." % prog["corePrinciple"]))
    a(("blank", ""))

    a(("body", "2. 모듈 구성"))
    for m in data["modules"]:
        a(("body", "모듈%d %s (%s) : %s" % (m["no"], m["name"], m["range"], m["tagline"])))
        a(("body", "  %s" % m["intent"]))
    a(("blank", ""))

    a(("body", "3. AI 신호등 4단계"))
    a(("body", "계획서에 있던 3단계(허용·조건부·제한)는 옛 기준입니다. 4단계를 씁니다."))
    for s in data["signals"]:
        a(("body", "%s (%s) : %s" % (s["light"], s["policy"], s["student"])))
        a(("body", "  %s" % s["meaning"]))
        a(("body", "  예) %s" % " / ".join(s["examples"])))
    a(("blank", ""))

    a(("body", "4. 우리 반 AI 약속 8조항의 뿌리"))
    a(("body", "7차시에 학생이 직접 만듭니다. 교사가 완성된 조항을 먼저 주지 않습니다."))
    for c in data["aiComponents"]:
        a(("body", "%s %s → %s (%d차시)" % (c["mark"], c["name"], c["pledge"], c["lesson"])))
    a(("blank", ""))

    a(("body", "5. 차시별 지도 유의점"))
    for l in data["lessons"]:
        a(("body", "%d차시 %s" % (l["no"], l["shortTitle"])))
        for c in l["cautions"]:
            a(("body", "  · %s" % c))
        a(("body", "  기기가 없을 때 : %s" % l["alternative"]))
        a(("body", "  웹앱 : %s" % l["webapp"]["name"]))
    a(("blank", ""))

    a(("body", "6. 부적정 활용 신호와 대응"))
    for sig, how in [
        ("학생이 결과물을 스스로 설명하지 못한다", "설명할 수 있으면 내 것이라는 기준을 다시 세우고 3단계 글쓰기로 돌아갑니다."),
        ("모르는 것을 생각해 보기 전에 AI부터 연다", "9차시 1단계를 먼저 채우게 하고, 채우기 전에는 다음 단계를 열지 않습니다."),
        ("친구 이름이나 사진을 넣으려 한다", "4차시 정보 분류 카드로 돌아가 왜 위험한지 근거를 말하게 합니다."),
        ("힘든 마음을 AI에게만 털어놓는다", "10차시 안내 문구를 함께 읽고, 필요하면 수업 후 개별 면담으로 연결합니다."),
        ("AI가 만든 것을 자기가 한 것처럼 낸다", "8차시 AI 활용 표기표를 다시 쓰게 합니다. 밝히는 것이 정직한 방법임을 강조합니다."),
    ]:
        a(("body", "· %s" % sig))
        a(("body", "  → %s" % how))
    a(("blank", ""))

    a(("body", "7. 도구 준비"))
    a(("body", "웹앱 12개는 홈페이지에서 바로 열립니다. 선생님이 방 코드를 만들어 화면에 띄워 주세요."))
    a(("body", "방 코드 6자리와 비밀번호 4자리로 들어갑니다. 실명은 쓰지 않고 닉네임만 씁니다."))
    a(("body", "기기가 부족하면 카드 교구 3종과 활동지로 같은 활동을 할 수 있습니다."))
    a(("body", "생성형 이미지 도구는 교사 계정으로만 씁니다."))
    a(("blank", ""))

    a(("body", "8. 평가"))
    for p in data["assessmentPlan"]:
        a(("body", "%s (%s)" % (p["skill"], ", ".join("%d차시" % n for n in p["lessons"]))))
        a(("body", "  목표 : %s" % p["goal"]))
        a(("body", "  상 : %s" % p["high"]))
        a(("body", "  중 : %s" % p["mid"]))
        a(("body", "  하 : %s" % p["low"]))
    a(("body", "하 수준 학생은 모둠 내 또래 설명과 교사 추가 발문으로 즉시 피드백합니다."))
    a(("blank", ""))

    a(("body", "9. 자주 묻는 질문"))
    for q, ans in [
        ("1인 1기기가 없어도 되나요", "됩니다. 모든 차시에 교사 시연, 모둠 공용, 인쇄 카드 대안을 함께 두었습니다."),
        ("생성형 AI 계정이 없어도 되나요", "됩니다. 2·3·8·9차시는 교사 시연 결과를 함께 검토하는 방식으로 진행할 수 있습니다."),
        ("학생 개인정보가 수집되나요", "수집하지 않습니다. 닉네임만 씁니다. 기록은 학년도가 끝나면 지웁니다."),
        ("시수를 더 확보해야 하나요", "아닙니다. 실과 6차시, 도덕 2차시, 국어 1차시 통합, 학교자율시간 또는 창의적 체험활동 4차시로 운영합니다."),
        ("순서를 바꿔도 되나요", "모듈 순서는 바꾸지 않기를 권합니다. 겪은 경험 위에 판단을 세우도록 설계했기 때문입니다."),
        ("자료를 고쳐 써도 되나요", "됩니다. CC BY-NC-SA로 공개합니다. 웹앱은 제작 프롬프트를 함께 두었으니 각자 고쳐 만드실 수 있습니다."),
    ]:
        a(("body", "Q. %s" % q))
        a(("body", "A. %s" % ans))
    a(("blank", ""))
    a(("body", prog["copyrightLine"]))

    out = os.path.join(T.ROOT, "out", "해설서", "WISE_교사용_해설서.hwpx")
    n = build_hwpx(L, out)
    print("만들었다 : %s  (문단 %d개)" % (out, n))
    return out


def main():
    data = T.load_lessons()
    what = sys.argv[1] if len(sys.argv) > 1 else "all"
    if what in ("진도표", "all"):
        make_progress(data)
    if what in ("교구", "all"):
        make_cards(data)
    if what in ("해설서", "all"):
        make_guide(data)
    return 0


if __name__ == "__main__":
    sys.exit(main())
