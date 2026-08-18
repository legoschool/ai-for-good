# -*- coding: utf-8 -*-
"""학교자율시간 진도표 XLSX 검증."""
import os
import sys

from openpyxl import load_workbook

import tasks as T

T.setup_console()

ERRORS = []
PATH = os.path.join(T.ROOT, "out", "서류", "WISE_학교자율시간_진도표.xlsx")


def err(m):
    ERRORS.append(m)


def cells(ws):
    out = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None:
                out.append(str(v))
    return "\n".join(out)


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else PATH
    if not os.path.isabs(path):
        path = os.path.join(T.ROOT, path)
    if not os.path.exists(path):
        print("실패  파일이 없다 : %s" % path)
        return 1

    wb = load_workbook(path)
    data = T.load_lessons()

    for need in ["진도표", "시수 배분", "평가 계획", "사전사후 설문"]:
        if need not in wb.sheetnames:
            err("시트가 없다 : %s" % need)

    if "진도표" in wb.sheetnames:
        ws = wb["진도표"]
        if ws.max_row != 13:
            err("진도표는 머리글 1행 + 12차시여야 한다 (현재 %d행)" % ws.max_row)
        text = cells(ws)
        for l in data["lessons"]:
            if l["shortTitle"] not in text:
                err("진도표에 %d차시가 없다" % l["no"])
            if l["webapp"]["name"] not in text:
                err("진도표에 %d차시 웹앱이 없다" % l["no"])
        if ws.freeze_panes != "A2":
            err("진도표 첫 행이 고정되지 않았다")

    if "시수 배분" in wb.sheetnames:
        ws = wb["시수 배분"]
        total = None
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "합계":
                total = row[2]
        if total != 13:
            err("시수 합계가 %s 다. 실과6 + 도덕2 + 국어1 + 자율4 = 13 이어야 한다" % total)
        text = cells(ws)
        if "7·8·10·12차시" not in text:
            err("학교자율시간 4차시(7·8·10·12)가 명시되지 않았다")

    if "평가 계획" in wb.sheetnames:
        ws = wb["평가 계획"]
        if ws.max_row != 8:
            err("평가 계획은 머리글 1행 + 중점 7역량이어야 한다 (현재 %d행)" % ws.max_row)

    if "사전사후 설문" in wb.sheetnames:
        ws = wb["사전사후 설문"]
        text = cells(ws)
        for i in data["survey"]["items"]:
            if i["text"] not in text:
                err("설문 문항 %d번이 없다" % i["no"])
        if "=IF(" not in text:
            err("변화량 자동 계산 수식이 없다")
        if "역채점" not in text:
            err("1번 문항 역채점 안내가 없다")

    blob = "\n".join(cells(wb[n]) for n in wb.sheetnames)
    if "—" in blob:
        err("em dash(—) 를 쓰지 않는다")
    if "허용·조건부·제한" in blob:
        err("옛 3단계 신호등 표현이 남아 있다")

    if ERRORS:
        for e in ERRORS:
            print("실패  %s" % e)
        print("")
        print("NG  오류 %d건" % len(ERRORS))
        return 1

    print("OK  %s" % os.path.basename(path))
    print("    시트 %d개 (%s)" % (len(wb.sheetnames), ", ".join(wb.sheetnames)))
    print("    12차시 전체, 시수 합계 13, 중점 7역량, 설문 8문항 자동 계산 확인")
    return 0


if __name__ == "__main__":
    sys.exit(main())
